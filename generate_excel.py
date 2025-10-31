#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import os

def create_grade_excel():
    """创建包含班级成绩表的Excel文件"""
    
    # 确保目标目录存在
    output_path = "/Users/issuser/Documents/file-test/sample.xlsx"
    output_dir = os.path.dirname(output_path)
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    # 创建工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "班级成绩表"
    
    # 设置表头
    headers = ["学号", "姓名", "语文", "数学", "英语", "物理", "化学", "总分", "平均分"]
    ws.append(headers)
    
    # 设置表头样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 学生成绩数据
    students_data = [
        ["2024001", "张三", 92, 88, 95, 87, 90],
        ["2024002", "李四", 85, 91, 88, 92, 86],
        ["2024003", "王五", 78, 82, 85, 79, 83],
        ["2024004", "赵六", 95, 97, 93, 96, 94],
        ["2024005", "钱七", 88, 86, 90, 85, 89],
        ["2024006", "孙八", 91, 89, 87, 90, 88],
        ["2024007", "周九", 83, 85, 82, 84, 86],
        ["2024008", "吴十", 76, 79, 81, 77, 80],
        ["2024009", "郑十一", 94, 92, 96, 93, 95],
        ["2024010", "王十二", 89, 91, 88, 90, 87],
        ["2024011", "刘十三", 87, 84, 86, 88, 85],
        ["2024012", "陈十四", 92, 95, 91, 93, 94],
    ]
    
    # 填充数据并计算总分和平均分
    for student in students_data:
        student_id, name, chinese, math, english, physics, chemistry = student
        total = chinese + math + english + physics + chemistry
        average = round(total / 5, 2)
        ws.append([student_id, name, chinese, math, english, physics, chemistry, total, average])
    
    # 设置列宽
    column_widths = [12, 10, 8, 8, 8, 8, 8, 10, 10]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    
    # 设置数据单元格居中
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 添加总结行
    summary_row = ws.max_row + 2
    ws[f'A{summary_row}'] = "班级统计"
    ws[f'A{summary_row}'].font = Font(bold=True, size=11)
    
    # 计算各科平均分
    ws[f'A{summary_row + 1}'] = "各科平均分："
    subjects = ["语文", "数学", "英语", "物理", "化学"]
    for i, subject in enumerate(subjects, 3):
        col_letter = openpyxl.utils.get_column_letter(i)
        avg_formula = f'=AVERAGE({col_letter}2:{col_letter}{ws.max_row - 2})'
        ws[f'{col_letter}{summary_row + 1}'] = avg_formula
        ws[f'{col_letter}{summary_row + 1}'].number_format = '0.00'
    
    # 保存Excel文件
    wb.save(output_path)
    print(f"Excel文件已成功创建：{output_path}")
    print(f"共包含 {len(students_data)} 条学生成绩数据")

if __name__ == "__main__":
    create_grade_excel()
